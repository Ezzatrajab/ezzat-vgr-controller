"""
Datahämtningsfunktioner för Ezzat's Controlling System
RÄTT STRUKTUR baserat på faktiska datakällor
"""

import pandas as pd
import os
import glob
from datetime import datetime
import openpyxl


# ========================================
# MAPPING: Enhetsnamn i KPIer-fil
# ========================================

# Mapping mellan KST och enhetsnamn i KPIer Stor-GBG.xlsx
KST_TO_KPINAME = {
    # VC-enheter
    '102': 'Omtanken Vårdcentral Frölunda Torg',
    '103': 'Omtanken Vårdcentral Grimmered',
    '104': 'Omtanken Vårdcentral Majorna',
    '106': 'Omtanken Vårdcentral Landala',
    '107': 'Omtanken Vårdcentral Pedagogen Park',
    '108-109': 'Omtanken Vårdcentral Åby',  # Åby + Kållered
    '110': 'Omtanken Vårdcentral Kviberg',
    '111': 'Omtanken Vårdcentral Olskroken',
    '302-303': 'Kvarterskliniken Avenyn',  # Avenyn + Lorensberg
    '304': 'Kvarterskliniken Husaren',
    '015': 'Kvarterskliniken Karlastaden',
    '4020': 'Citysjukhuset Plus 7 vårdcentral',
}

# Mapping för kort namn (används i rad 322-336 för listning 2026)
KST_TO_SHORTNAME = {
    '102': 'Frölunda Torg',
    '103': 'Grimmered',
    '104': 'Majorna',
    '106': 'Landala',
    '107': 'Pedagogen Park',
    '108-109': 'Åby',  # OBS: Kallas "Åby" i listning-sektionen
    '110': 'Kviberg',
    '111': 'Olskroken',
    '302-303': 'Avenyn',
    '304': 'Husaren',
    '015': 'Karlastaden',
    '4020': 'City',
}


# ========================================
# FILSÖKVÄGAR
# ========================================

def get_file_paths(enhet_kst, base_path=None):
    """Returnerar sökvägar till alla datafiler för en enhet"""
    if base_path is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        base_path = os.path.join(script_dir, 'data')

    enhet_path = os.path.join(base_path, enhet_kst)

    # Olika filnamn för olika enheter (baserat på tidigare mapping)
    file_map = {
        # Rehab-enheter
        '601': {'fte_actual': 'FTE Producerande per Yrkesgrupp (7).xlsx', 'hr_cost': 'HR Cost (12).xlsx', 'intakt_budget': 'Intäkt Budget Rehab (26).xlsx', 'pl_actual': 'P&L Actual.xlsx', 'pl_budget': 'P&L Budget.xlsx'},
        '602': {'fte_actual': 'FTE Producerande per Yrkesgrupp (8).xlsx', 'hr_cost': 'HR Cost (13).xlsx', 'intakt_budget': 'Intäkt Budget Rehab (27).xlsx', 'pl_actual': 'P&L Actual.xlsx', 'pl_budget': 'P&L Budget.xlsx'},
        '603': {'fte_actual': 'FTE Producerande per Yrkesgrupp (19).xlsx', 'hr_cost': 'HR Cost (25).xlsx', 'intakt_budget': 'Intäkt Budget Rehab (28).xlsx', 'pl_actual': 'P&L Actual.xlsx', 'pl_budget': 'P&L Budget.xlsx'},
        '604': {'fte_actual': 'FTE Producerande per Yrkesgrupp (20).xlsx', 'hr_cost': 'HR Cost (26).xlsx', 'intakt_budget': 'Intäkt Budget Rehab (29).xlsx', 'pl_actual': 'P&L Actual.xlsx', 'pl_budget': 'P&L Budget.xlsx'},
        '605': {'fte_actual': 'FTE Producerande per Yrkesgrupp (21).xlsx', 'hr_cost': 'HR Cost (27).xlsx', 'intakt_budget': 'Intäkt Budget Rehab (30).xlsx', 'pl_actual': 'P&L Actual.xlsx', 'pl_budget': 'P&L Budget.xlsx'},
        '607': {'fte_actual': 'FTE Producerande per Yrkesgrupp (22).xlsx', 'hr_cost': 'HR Cost (28).xlsx', 'intakt_budget': 'Intäkt Budget Rehab (31).xlsx', 'pl_actual': 'P&L Actual.xlsx', 'pl_budget': 'P&L Budget.xlsx'},
        '660': {'fte_actual': 'FTE Producerande per Yrkesgrupp (23).xlsx', 'hr_cost': 'HR Cost (29).xlsx', 'intakt_budget': 'Intäkt Budget Rehab (32).xlsx', 'pl_actual': 'P&L Actual.xlsx', 'pl_budget': 'P&L Budget.xlsx'},
        '715': {'fte_actual': 'FTE Producerande per Yrkesgrupp (24).xlsx', 'hr_cost': 'HR Cost (30).xlsx', 'intakt_budget': 'Intäkt Budget Rehab (33).xlsx', 'pl_actual': 'P&L Actual.xlsx', 'pl_budget': 'P&L Budget.xlsx'},

        # VC-enheter
        '102': {'fte_actual': 'FTE Producerande per Yrkesgrupp (4).xlsx', 'hr_cost': 'HR Cost (10).xlsx', 'pl_actual': 'P&L Actual.xlsx', 'pl_budget': 'P&L Budget.xlsx'},
        '103': {'fte_actual': 'FTE Producerande per Yrkesgrupp (5).xlsx', 'hr_cost': 'HR Cost (11).xlsx', 'pl_actual': 'P&L Actual.xlsx', 'pl_budget': 'P&L Budget.xlsx'},
        '104': {'fte_actual': 'FTE Producerande per Yrkesgrupp (9).xlsx', 'hr_cost': 'HR Cost (14).xlsx', 'pl_actual': 'P&L Actual.xlsx', 'pl_budget': 'P&L Budget.xlsx'},
        '106': {'fte_actual': 'FTE Producerande per Yrkesgrupp (10).xlsx', 'hr_cost': 'HR Cost (15).xlsx', 'pl_actual': 'P&L Actual.xlsx', 'pl_budget': 'P&L Budget.xlsx'},
        '107': {'fte_actual': 'FTE Producerande per Yrkesgrupp (11).xlsx', 'hr_cost': 'HR Cost (16).xlsx', 'pl_actual': 'P&L Actual.xlsx', 'pl_budget': 'P&L Budget.xlsx'},
        '108-109': {'fte_actual': 'FTE Producerande per Yrkesgrupp (12).xlsx', 'hr_cost': 'HR Cost (17).xlsx', 'pl_actual': 'P&L Actual.xlsx', 'pl_budget': 'P&L Budget.xlsx'},
        '110': {'fte_actual': 'FTE Producerande per Yrkesgrupp (13).xlsx', 'hr_cost': 'HR Cost (19).xlsx', 'pl_actual': 'P&L Actual.xlsx', 'pl_budget': 'P&L Budget.xlsx'},
        '111': {'fte_actual': 'FTE Producerande per Yrkesgrupp (14).xlsx', 'hr_cost': 'HR Cost (20).xlsx', 'pl_actual': 'P&L Actual.xlsx', 'pl_budget': 'P&L Budget.xlsx'},
        '302-303': {'fte_actual': 'FTE Producerande per Yrkesgrupp (15).xlsx', 'hr_cost': 'HR Cost (21).xlsx', 'pl_actual': 'P&L Actual.xlsx', 'pl_budget': 'P&L Budget.xlsx'},
        '304': {'fte_actual': 'FTE Producerande per Yrkesgrupp (16).xlsx', 'hr_cost': 'HR Cost (22).xlsx', 'pl_actual': 'P&L Actual.xlsx', 'pl_budget': 'P&L Budget.xlsx'},
        '015': {'fte_actual': 'FTE Producerande per Yrkesgrupp (17).xlsx', 'hr_cost': 'HR Cost (23).xlsx', 'pl_actual': 'P&L Actual.xlsx', 'pl_budget': 'P&L Budget.xlsx'},
        '4020': {'fte_actual': 'FTE Producerande per Yrkesgrupp (18).xlsx', 'hr_cost': 'HR Cost (24).xlsx', 'pl_actual': 'P&L Actual.xlsx', 'pl_budget': 'P&L Budget.xlsx'},
    }

    if enhet_kst not in file_map:
        raise ValueError(f"Enhet {enhet_kst} finns inte i file_map")

    paths = {}
    for key, filename in file_map[enhet_kst].items():
        paths[key] = os.path.join(enhet_path, filename)

    return paths


# ========================================
# DATAHÄMTNING FRÅN KPIer Stor-GBG.xlsx
# ========================================

def load_kpi_data_from_file(enhet_kst, manad_str, base_path=None):
    """
    Hämtar Listning Actual och ACG Casemix Actual från KPIer Stor-GBG.xlsx

    OBS: Slår ihop data för kombinerade enheter:
    - 108-109 (Åby-Kållered): Åby + Kållered
    - 302-303 (Avenyn-Lorensberg): Avenyn + Lorensberg

    Args:
        enhet_kst: '102', '103', '015', etc
        manad_str: '2026-01', '2026-02', etc
        base_path: Bas-sökväg till data-mappen

    Returns:
        dict: {
            'listning_actual': int,
            'acg_casemix_actual': float
        }
    """
    try:
        if base_path is None:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            base_path = os.path.join(script_dir, 'data')

        kpi_file = os.path.join(base_path, 'KPIer Stor-GBG.xlsx')

        if not os.path.exists(kpi_file):
            print(f"KPIer-fil hittades inte: {kpi_file}")
            return {'listning_actual': 0, 'acg_casemix_actual': 0}

        wb = openpyxl.load_workbook(kpi_file, data_only=True)
        ws = wb['Data']

        # Konvertera månad till period-format (202601, 202602)
        year, month = manad_str.split('-')
        period = int(year + month)

        # LISTNING ACTUAL: Rad 322-336, kolumner 3-4 för 202601-202602

        listning_actual = 0

        # Hitta kolumn-index för rätt period
        header_row_idx = 322
        col_idx = None
        for col in range(1, 10):
            val = ws.cell(header_row_idx, col).value
            if val and str(val) == str(period):
                col_idx = col
                break

        if col_idx:
            # Specialhantering för kombinerade enheter
            if enhet_kst == '108-109':
                # Åby + Kållered (Omtanken Vårdcentral Åby + Omtanken Vårdcentral Kållered)
                listning_aby = 0
                listning_kallered = 0
                for row_idx in range(323, 337):
                    cell_val = ws.cell(row_idx, 1).value
                    if cell_val:
                        cell_str = str(cell_val).strip()
                        # Åby (men inte Kållered)
                        if 'Åby' in cell_str and 'Kållered' not in cell_str:
                            listning_val = ws.cell(row_idx, col_idx).value
                            listning_aby = int(listning_val) if pd.notna(listning_val) else 0
                        # Kållered
                        elif 'Kållered' in cell_str:
                            listning_val = ws.cell(row_idx, col_idx).value
                            listning_kallered = int(listning_val) if pd.notna(listning_val) else 0
                listning_actual = listning_aby + listning_kallered

            elif enhet_kst == '302-303':
                # Avenyn + Lorensberg (Kvarterskliniken Avenyn + Kvarterskliniken Lorensberg)
                listning_avenyn = 0
                listning_lorensberg = 0
                for row_idx in range(323, 337):
                    cell_val = ws.cell(row_idx, 1).value
                    if cell_val:
                        cell_str = str(cell_val).strip()
                        # Kvarterskliniken Avenyn (men inte Lorensberg)
                        if 'Avenyn' in cell_str and 'Lorensberg' not in cell_str:
                            listning_val = ws.cell(row_idx, col_idx).value
                            listning_avenyn = int(listning_val) if pd.notna(listning_val) else 0
                        # Kvarterskliniken Lorensberg
                        elif 'Lorensberg' in cell_str:
                            listning_val = ws.cell(row_idx, col_idx).value
                            listning_lorensberg = int(listning_val) if pd.notna(listning_val) else 0
                listning_actual = listning_avenyn + listning_lorensberg

            else:
                # Normal enhet
                short_name = KST_TO_SHORTNAME.get(enhet_kst)
                if short_name:
                    for row_idx in range(323, 337):
                        cell_val = ws.cell(row_idx, 1).value
                        if cell_val and short_name in str(cell_val):
                            listning_val = ws.cell(row_idx, col_idx).value
                            listning_actual = int(listning_val) if pd.notna(listning_val) else 0
                            break

        # ACG CASEMIX ACTUAL: Rad 153-168, kolumn 27-28 för 202601-202602

        acg_casemix_actual = 0
        casemix_col = 27 if period == 202601 else (28 if period == 202602 else None)

        if casemix_col:
            # Specialhantering för kombinerade enheter (vägt genomsnitt)
            if enhet_kst == '108-109':
                # Åby + Kållered (Omtanken Vårdcentral Åby + Omtanken Vårdcentral Kållered)
                casemix_aby = 0
                casemix_kallered = 0
                listning_aby_casemix = 0
                listning_kallered_casemix = 0

                for row_idx in range(154, 169):
                    cell_val = ws.cell(row_idx, 1).value
                    if cell_val:
                        cell_str = str(cell_val).strip()
                        # Åby (men inte Kållered)
                        if 'Åby' in cell_str and 'Kållered' not in cell_str:
                            casemix_val = ws.cell(row_idx, casemix_col).value
                            casemix_aby = float(casemix_val) if pd.notna(casemix_val) else 0
                            # Hämta listning från rad 198-212
                            for list_row in range(199, 213):
                                list_cell = ws.cell(list_row, 1).value
                                if list_cell:
                                    list_cell_str = str(list_cell).strip()
                                    if 'Åby' in list_cell_str and 'Kållered' not in list_cell_str:
                                        list_val = ws.cell(list_row, casemix_col).value
                                        listning_aby_casemix = int(list_val) if pd.notna(list_val) else 0
                                        break
                        # Kållered
                        elif 'Kållered' in cell_str:
                            casemix_val = ws.cell(row_idx, casemix_col).value
                            casemix_kallered = float(casemix_val) if pd.notna(casemix_val) else 0
                            for list_row in range(199, 213):
                                list_cell = ws.cell(list_row, 1).value
                                if list_cell and 'Kållered' in str(list_cell):
                                    list_val = ws.cell(list_row, casemix_col).value
                                    listning_kallered_casemix = int(list_val) if pd.notna(list_val) else 0
                                    break

                # Beräkna vägt genomsnitt
                total_listning = listning_aby_casemix + listning_kallered_casemix
                if total_listning > 0:
                    acg_casemix_actual = (casemix_aby * listning_aby_casemix + casemix_kallered * listning_kallered_casemix) / total_listning

            elif enhet_kst == '302-303':
                # Avenyn + Lorensberg (Kvarterskliniken Avenyn + Kvarterskliniken Lorensberg)
                casemix_avenyn = 0
                casemix_lorensberg = 0
                listning_avenyn_casemix = 0
                listning_lorensberg_casemix = 0

                for row_idx in range(154, 169):
                    cell_val = ws.cell(row_idx, 1).value
                    if cell_val:
                        cell_str = str(cell_val).strip()
                        # Kvarterskliniken Avenyn (men inte Lorensberg)
                        if 'Avenyn' in cell_str and 'Lorensberg' not in cell_str:
                            casemix_val = ws.cell(row_idx, casemix_col).value
                            casemix_avenyn = float(casemix_val) if pd.notna(casemix_val) else 0
                            for list_row in range(199, 213):
                                list_cell = ws.cell(list_row, 1).value
                                if list_cell:
                                    list_cell_str = str(list_cell).strip()
                                    if 'Avenyn' in list_cell_str and 'Lorensberg' not in list_cell_str:
                                        list_val = ws.cell(list_row, casemix_col).value
                                        listning_avenyn_casemix = int(list_val) if pd.notna(list_val) else 0
                                        break
                        # Kvarterskliniken Lorensberg
                        elif 'Lorensberg' in cell_str:
                            casemix_val = ws.cell(row_idx, casemix_col).value
                            casemix_lorensberg = float(casemix_val) if pd.notna(casemix_val) else 0
                            for list_row in range(199, 213):
                                list_cell = ws.cell(list_row, 1).value
                                if list_cell and 'Lorensberg' in str(list_cell):
                                    list_val = ws.cell(list_row, casemix_col).value
                                    listning_lorensberg_casemix = int(list_val) if pd.notna(list_val) else 0
                                    break

                # Beräkna vägt genomsnitt
                total_listning = listning_avenyn_casemix + listning_lorensberg_casemix
                if total_listning > 0:
                    acg_casemix_actual = (casemix_avenyn * listning_avenyn_casemix + casemix_lorensberg * listning_lorensberg_casemix) / total_listning

            else:
                # Normal enhet
                full_name = KST_TO_KPINAME.get(enhet_kst)
                if full_name:
                    for row_idx in range(154, 169):
                        cell_val = ws.cell(row_idx, 1).value
                        if cell_val and full_name in str(cell_val):
                            casemix_val = ws.cell(row_idx, casemix_col).value
                            acg_casemix_actual = float(casemix_val) if pd.notna(casemix_val) else 0
                            break

        wb.close()

        return {
            'listning_actual': listning_actual,
            'acg_casemix_actual': acg_casemix_actual
        }

    except Exception as e:
        print(f"Fel vid läsning från KPIer-fil för {enhet_kst}, {manad_str}: {e}")
        return {'listning_actual': 0, 'acg_casemix_actual': 0}


# ========================================
# BUDGET FRÅN INTÄKT BUDGET VC
# ========================================

def load_vc_budget(enhet_kst, manad_str, base_path=None):
    """
    Hämtar budget-data för VC från "Intäkt Budget VC"

    Returns:
        dict: {
            'listning': int,
            'acg_casemix': float
        }
    """
    try:
        if base_path is None:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            base_path = os.path.join(script_dir, 'data')

        enhet_path = os.path.join(base_path, enhet_kst)

        # Hitta Intäkt Budget VC-filen
        vc_files = glob.glob(os.path.join(enhet_path, 'Intäkt Budget VC*.xlsx'))

        if not vc_files:
            print(f"Ingen Intäkt Budget VC-fil hittades för {enhet_kst}")
            return {'listning': 0, 'acg_casemix': 0}

        df = pd.read_excel(vc_files[0], header=None)

        # Konvertera månad till månad-namn
        year, month = manad_str.split('-')
        month_num = int(month)
        month_names = ['January', 'February', 'March', 'April', 'May', 'June',
                       'July', 'August', 'September', 'October', 'November', 'December']
        month_name = month_names[month_num - 1]

        # Hitta kolumn-index för månaden
        col_idx = None
        for i, val in enumerate(df.iloc[0]):
            if val == month_name:
                col_idx = i
                break

        if col_idx is None:
            print(f"Månad {month_name} hittades inte i budget-fil för {enhet_kst}")
            return {'listning': 0, 'acg_casemix': 0}

        # Läs värden:
        # Rad 8 (index 7): ACG casemix % (Vårdtyngd)
        # Rad 9 (index 8): Antal listade patienter

        acg_casemix = df.iloc[7, col_idx] if pd.notna(df.iloc[7, col_idx]) else 0
        listning = df.iloc[8, col_idx] if pd.notna(df.iloc[8, col_idx]) else 0

        return {
            'listning': int(listning),
            'acg_casemix': float(acg_casemix)
        }

    except Exception as e:
        print(f"Fel vid läsning av VC-budget för {enhet_kst}, {manad_str}: {e}")
        return {'listning': 0, 'acg_casemix': 0}


# ========================================
# FTE, PERSONALKOSTNAD (samma som tidigare)
# ========================================

def load_fte_actual(enhet_kst, manad_str, base_path=None):
    """Hämtar actual FTE från FTE Producerande per Yrkesgrupp"""
    try:
        paths = get_file_paths(enhet_kst, base_path)
        df = pd.read_excel(paths['fte_actual'])

        year, month = manad_str.split('-')
        period = int(year + month)

        for idx, row in df.iterrows():
            if idx == 0:
                continue
            if idx > 20:
                break

            period_val = row['Yrkesgrupp']
            if pd.notna(period_val) and isinstance(period_val, (int, float)) and int(period_val) == period:
                return float(row['Total']) if pd.notna(row['Total']) else 0.0

        return 0.0

    except Exception as e:
        print(f"Fel vid läsning av FTE actual för {enhet_kst}, {manad_str}: {e}")
        return 0.0


def load_fte_budget(enhet_kst, manad_str, base_path=None):
    """Hämtar budget FTE från HR Cost"""
    try:
        paths = get_file_paths(enhet_kst, base_path)
        df = pd.read_excel(paths['hr_cost'])

        year, month = manad_str.split('-')
        month_num = int(month)
        month_names = ['January', 'February', 'March', 'April', 'May', 'June',
                       'July', 'August', 'September', 'October', 'November', 'December']
        month_name = month_names[month_num - 1]

        fte_col = f"{month_name}.1"

        if fte_col not in df.columns:
            print(f"Kolumn {fte_col} finns inte i HR Cost för {enhet_kst}")
            return 0.0

        total_rows = df[df['Month'] == 'Total']

        if total_rows.empty:
            print(f"Ingen Total-rad hittades i HR Cost för {enhet_kst}")
            return 0.0

        total_row = total_rows.iloc[0]
        fte_budget = total_row[fte_col]

        return float(fte_budget) if pd.notna(fte_budget) else 0.0

    except Exception as e:
        print(f"Fel vid läsning av FTE budget för {enhet_kst}, {manad_str}: {e}")
        return 0.0


def load_rehab_poang_budget(enhet_kst, manad_str, base_path=None):
    """
    Hämtar budgeterade Rehab-poäng från "Intäkt Budget Rehab"
    OBS: Endast för Rehab-enheter (601, 602, 603, 604, 605, 607, 660, 715)

    Args:
        enhet_kst: '601', '602', etc
        manad_str: '2026-01', '2026-02' etc
        base_path: Bas-sökväg till data-mappen

    Returns:
        dict: {
            'maaltal': float (Måltal per prestationsanställd),
            'antal_anstallda': float (Antal Prestationsanställda),
            'budgeterad_intakt': float (Intäkt konto 3053)
        }
    """
    try:
        # Endast för Rehab-enheter
        if enhet_kst not in ['601', '602', '603', '604', '605', '607', '660', '715']:
            return {'maaltal': 0, 'antal_anstallda': 0, 'budgeterad_intakt': 0}

        paths = get_file_paths(enhet_kst, base_path)
        df = pd.read_excel(paths['intakt_budget'])

        # Konvertera månad till månad-namn
        year, month = manad_str.split('-')
        month_num = int(month)
        month_names = ['January', 'February', 'March', 'April', 'May', 'June',
                       'July', 'August', 'September', 'October', 'November', 'December']
        month_name = month_names[month_num - 1]

        if month_name not in df.columns:
            print(f"Kolumn {month_name} finns inte i Intäkt Budget Rehab för {enhet_kst}")
            return {'maaltal': 0, 'antal_anstallda': 0, 'budgeterad_intakt': 0}

        # Rad 2 = Måltal per prestationsanställd
        # Rad 3 = Antal Prestationsanställda
        # Rad 4 = Intäkt konto 3053

        maaltal = df.iloc[2][month_name] if pd.notna(df.iloc[2][month_name]) else 0
        antal_anstallda = df.iloc[3][month_name] if pd.notna(df.iloc[3][month_name]) else 0
        budgeterad_intakt = df.iloc[4][month_name] if pd.notna(df.iloc[4][month_name]) else 0

        return {
            'maaltal': float(maaltal),
            'antal_anstallda': float(antal_anstallda),
            'budgeterad_intakt': float(budgeterad_intakt)
        }

    except Exception as e:
        print(f"Fel vid läsning av Rehab poäng budget för {enhet_kst}, {manad_str}: {e}")
        return {'maaltal': 0, 'antal_anstallda': 0, 'budgeterad_intakt': 0}


def load_personalkostnad(enhet_kst, manad_str, base_path=None):
    """Hämtar personalkostnader från P&L Actual och P&L Budget"""
    try:
        paths = get_file_paths(enhet_kst, base_path)

        df_actual = pd.read_excel(paths['pl_actual'])
        df_budget = pd.read_excel(paths['pl_budget'])

        year, month = manad_str.split('-')
        manad_num = int(year) * 100 + int(month)

        # Hitta kolumn för rätt månad
        period_cols = [col for col in df_actual.columns if 'Selected Period' in str(col)]

        header_row_actual = df_actual.iloc[0]
        header_row_budget = df_budget.iloc[0]

        col_idx = None
        for col in period_cols:
            val = header_row_actual[col]
            if pd.notna(val) and int(val) == manad_num:
                col_idx = col
                break

        if col_idx is None:
            return {'actual': 0, 'budget': 0}

        # Hitta Medical staff Total
        medical_staff_actual = df_actual[
            (df_actual['CG Item'] == 'COGS') &
            (df_actual['Unnamed: 1'] == 'Medical staff') &
            (df_actual['Unnamed: 2'] == 'Total')
        ]

        medical_staff_budget = df_budget[
            (df_budget['CG Item'] == 'COGS') &
            (df_budget['Unnamed: 1'] == 'Medical staff') &
            (df_budget['Unnamed: 2'] == 'Total')
        ]

        if medical_staff_actual.empty or medical_staff_budget.empty:
            # Fallback: COGS Total
            cogs_total_actual = df_actual[
                (df_actual['CG Item'] == 'COGS') &
                (df_actual['Unnamed: 1'] == 'Total')
            ]
            cogs_total_budget = df_budget[
                (df_budget['CG Item'] == 'COGS') &
                (df_budget['Unnamed: 1'] == 'Total')
            ]

            if cogs_total_actual.empty or cogs_total_budget.empty:
                return {'actual': 0, 'budget': 0}

            actual_val = cogs_total_actual.iloc[0][col_idx]
            budget_val = cogs_total_budget.iloc[0][col_idx]
        else:
            actual_val = medical_staff_actual.iloc[0][col_idx]
            budget_val = medical_staff_budget.iloc[0][col_idx]

        return {
            'actual': abs(float(actual_val)) if pd.notna(actual_val) else 0,
            'budget': abs(float(budget_val)) if pd.notna(budget_val) else 0
        }

    except Exception as e:
        print(f"Fel vid läsning av personalkostnader för {enhet_kst}, {manad_str}: {e}")
        return {'actual': 0, 'budget': 0}


# ========================================
# HUVUDFUNKTION: Hämta all data
# ========================================

def load_all_data_for_enhet(enhet_kst, manad_str, base_path=None):
    """
    Hämtar ALL data för en enhet och månad

    Returns:
        dict: {
            'fte_actual': float,
            'fte_budget': float,
            'personalkostnad_actual': float,
            'personalkostnad_budget': float,
            'listning_actual': int,
            'listning_budget': int,
            'acg_casemix_actual': float,
            'acg_casemix_budget': float,
            'rehab_budget': dict (endast för Rehab-enheter)
        }
    """
    # FTE och Personalkostnad (samma för alla enheter)
    fte_actual = load_fte_actual(enhet_kst, manad_str, base_path)
    fte_budget = load_fte_budget(enhet_kst, manad_str, base_path)

    personalkostnad = load_personalkostnad(enhet_kst, manad_str, base_path)

    # Bas-data som returneras
    data = {
        'fte_actual': fte_actual,
        'fte_budget': fte_budget,
        'personalkostnad_actual': personalkostnad['actual'],
        'personalkostnad_budget': personalkostnad['budget'],
    }

    # Kolla om det är en Rehab-enhet
    if enhet_kst in ['601', '602', '603', '604', '605', '607', '660', '715']:
        # Rehab-enhet: Lägg till Rehab-budget
        data['rehab_budget'] = load_rehab_poang_budget(enhet_kst, manad_str, base_path)
        # Sätt 0 för VC-relaterade KPIer
        data['listning_actual'] = 0
        data['listning_budget'] = 0
        data['acg_casemix_actual'] = 0
        data['acg_casemix_budget'] = 0
    else:
        # VC-enhet: Hämta listning och casemix från KPIer och Budget-filer
        kpi_data = load_kpi_data_from_file(enhet_kst, manad_str, base_path)
        budget_data = load_vc_budget(enhet_kst, manad_str, base_path)

        data['listning_actual'] = kpi_data['listning_actual']
        data['listning_budget'] = budget_data['listning']
        data['acg_casemix_actual'] = kpi_data['acg_casemix_actual']
        data['acg_casemix_budget'] = budget_data['acg_casemix']

    return data


# ========================================
# TEST
# ========================================

if __name__ == "__main__":
    print("=" * 80)
    print("TEST: Landala (KST 106), Januari 2026")
    print("=" * 80)

    data = load_all_data_for_enhet('106', '2026-01')

    print(f"\nListning Actual: {data['listning_actual']:,}")
    print(f"Listning Budget: {data['listning_budget']:,}")
    print(f"ACG Casemix Actual: {data['acg_casemix_actual']:.3f}")
    print(f"ACG Casemix Budget: {data['acg_casemix_budget']:.2f}")
    print(f"FTE Actual: {data['fte_actual']:.2f}")
    print(f"FTE Budget: {data['fte_budget']:.2f}")
    print(f"Personalkostnad Actual: {data['personalkostnad_actual']:,.0f} kr")
    print(f"Personalkostnad Budget: {data['personalkostnad_budget']:,.0f} kr")

    print("\n" + "=" * 80)
    print("TEST: Karlastaden (KST 015), Februari 2026")
    print("=" * 80)

    data = load_all_data_for_enhet('015', '2026-02')

    print(f"\nListning Actual: {data['listning_actual']:,}")
    print(f"Listning Budget: {data['listning_budget']:,}")
    print(f"ACG Casemix Actual: {data['acg_casemix_actual']:.3f}")
    print(f"ACG Casemix Budget: {data['acg_casemix_budget']:.2f}")
    print(f"FTE Actual: {data['fte_actual']:.2f}")
    print(f"FTE Budget: {data['fte_budget']:.2f}")
    print(f"Personalkostnad Actual: {data['personalkostnad_actual']:,.0f} kr")
    print(f"Personalkostnad Budget: {data['personalkostnad_budget']:,.0f} kr")
